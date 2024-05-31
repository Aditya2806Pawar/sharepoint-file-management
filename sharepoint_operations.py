import io
import pandas as pd
import joblib
from openpyxl import Workbook, load_workbook
from office365.sharepoint.files.file import File
from authentication import CTX

def direct_upload_to_sharepoint_as_df(df, filename, target_url):
    writer_obj = io.BytesIO()
    df.to_excel(writer_obj, index=False)
    writer_obj.seek(0)
    target_folder = CTX.web.get_folder_by_server_relative_url(target_url)
    target_file = target_folder.upload_file(filename, writer_obj)
    CTX.execute_query()
    print(filename, "has been uploaded to url:", target_file.serverRelativeUrl)

def download_from_sharepoint_as_df(file_url):
    response = File.open_binary(CTX, file_url)
    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)
    df = pd.read_excel(bytes_file_obj, engine='openpyxl')
    return df

def direct_upload_to_sharepoint_as_pickle(pickle_obj, filename, target_url):
    bytes_obj = io.BytesIO()
    joblib.dump(pickle_obj, bytes_obj)
    bytes_obj.seek(0)
    target_folder = CTX.web.get_folder_by_server_relative_url(target_url)
    target_file = target_folder.upload_file(filename, bytes_obj)
    CTX.execute_query()
    print(filename, "has been uploaded to url:", target_file.serverRelativeUrl)

def download_from_sharepoint_as_pickle(file_url):
    response = File.open_binary(CTX, file_url)
    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)
    pickle_obj = joblib.load(bytes_file_obj)
    return pickle_obj

def direct_upload_to_sharepoint_as_workbook(workbook, filename, target_url):
    bytes_obj = io.BytesIO()
    workbook.save(bytes_obj)
    bytes_obj.seek(0)
    target_folder = CTX.web.get_folder_by_server_relative_url(target_url)
    target_file = target_folder.upload_file(filename, bytes_obj)
    CTX.execute_query()
    print(filename, "has been uploaded to url:", target_file.serverRelativeUrl)

def download_from_sharepoint_as_workbook(file_url):
    response = File.open_binary(CTX, file_url)
    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)
    workbook = load_workbook(bytes_file_obj)
    return workbook

def direct_upload_to_sharepoint_as_csv(df, filename, target_url):
    csv_obj = io.StringIO()
    df.to_csv(csv_obj, index=False)
    csv_obj.seek(0)
    bytes_obj = io.BytesIO(csv_obj.getvalue().encode())
    target_folder = CTX.web.get_folder_by_server_relative_url(target_url)
    target_file = target_folder.upload_file(filename, bytes_obj)
    CTX.execute_query()
    print(filename, "has been uploaded to url:", target_file.serverRelativeUrl)

def download_from_sharepoint_as_csv(file_url):
    response = File.open_binary(CTX, file_url)
    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)
    csv_obj = io.StringIO(bytes_file_obj.getvalue().decode())
    df = pd.read_csv(csv_obj)
    return df

# Example usage
if __name__ == "__main__":
    df = pd.DataFrame({'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']})
    target_url = "/sites/your_site_name/Shared Documents/your_folder_name"
    filename_excel = "example.xlsx"
    filename_pickle = "example.pkl"
    filename_csv = "example.csv"

    # Upload and download Excel
    direct_upload_to_sharepoint_as_df(df, filename_excel, target_url)
    df_downloaded = download_from_sharepoint_as_df(f"{target_url}/{filename_excel}")
    print(df_downloaded)

    # Upload and download Pickle
    pickle_obj = {'key': 'value'}
    direct_upload_to_sharepoint_as_pickle(pickle_obj, filename_pickle, target_url)
    pickle_downloaded = download_from_sharepoint_as_pickle(f"{target_url}/{filename_pickle}")
    print(pickle_downloaded)

    # Upload and download CSV
    direct_upload_to_sharepoint_as_csv(df, filename_csv, target_url)
    df_downloaded_csv = download_from_sharepoint_as_csv(f"{target_url}/{filename_csv}")
    print(df_downloaded_csv)

    # Upload and download openpyxl Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hello"
    sheet["B1"] = "World"
    filename_workbook = "example_workbook.xlsx"
    direct_upload_to_sharepoint_as_workbook(workbook, filename_workbook, target_url)
    workbook_downloaded = download_from_sharepoint_as_workbook(f"{target_url}/{filename_workbook}")
    sheet_downloaded = workbook_downloaded.active
    print(sheet_downloaded["A1"].value, sheet_downloaded["B1"].value)
